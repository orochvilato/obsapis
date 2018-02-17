# -*- coding: utf-8 -*-

from obsapis import mdb

def getScrutinsCles():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import markdown
    import requests

    r =requests.get("https://docs.google.com/spreadsheets/d/1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw/export?format=xlsx&id=1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw")
    f = StringIO(r.content)
    wb = load_workbook(f)
    scles = {}

    for n in range(2):
        ws = wb[wb.sheetnames[n]]
        elts = []
        for j,row in enumerate(ws.iter_rows(min_row=2)):
            elts.append([a.value for a in row])
        for e in elts:
            if e[2] != None:
                #scles[int(e[0])] = dict(num=int(e[0]),nom=e[1],descfmt=markdown.markdown(e[2]),desc=e[2].replace('_','').replace('*',''),theme=e[3],lien=e[4],lien_texte=e[5],lien_source=e[6],inversion=e[8].strip().lower(),dossier=e[7],niveau=n+1)
                scles[int(e[0])] = dict(num=int(e[0]),nom=e[1],descfmt=markdown.markdown(e[2]),desc=e[2],theme=e[3],lien=e[4],lien_texte=e[5],lien_source=e[6],inversion=e[8].strip().lower(),dossier=e[7],niveau=n+1)
    return scles

def getScrutinsPositions():
    spositions = {}
    for s in mdb.scrutins.find({},{'scrutin_id':1,'scrutin_num':1,'_id':None,'scrutin_positions.assemblee':1}):
        spositions[s['scrutin_num']] = s['scrutin_positions']['assemblee']
    return spositions
def getScrutinsSorts():
    ssorts = {}
    for s in mdb.scrutins.find({},{'scrutin_id':1,'scrutin_num':1,'_id':None,'scrutin_sort':1}):
        ssorts[s['scrutin_num']] = s['scrutin_sort']
    return ssorts
def getScrutinsData():
    sdata = {}
    for s in mdb.scrutins.find({},{'scrutin_id':1,'scrutin_num':1,'_id':None,'scrutin_sort':1,'scrutin_urlAmendement':1,'scrutin_lientexte':1}):
        sdata[s['scrutin_num']] = {'sort':s['scrutin_sort'],'urlAmendement':s.get('scrutin_urlAmendement',None),'scrutin_lientexte':s.get('scrutin_lientexte',[])}
    return sdata
