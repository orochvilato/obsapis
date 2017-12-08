# -*- coding: utf-8 -*-

import requests
import re
import unicodedata
def strip_accents(s):
   return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
def normalize(s):
    return strip_accents(s).replace(u'\u2019','').replace('&apos;','').replace(u'\xa0','').encode('utf8').replace(' ','').replace("'",'').replace('-','').replace('\x0a','').replace('\xc5\x93','oe').decode('utf8').lower() if s else s


def import_obsgouv_gdoc():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import requests

    r =requests.get("https://docs.google.com/spreadsheets/d/1mvBb66ulZJmdgWtnIgX9xpSwn8Xclv7p0h7rl1PQ_HA/export?format=xlsx&id=1mvBb66ulZJmdgWtnIgX9xpSwn8Xclv7p0h7rl1PQ_HA")
    f = StringIO(r.content)
    wb = load_workbook(f)


    entites = {}
    liens = []
    for sheet in (u'Collaborateurs des ministres',u'Collaborateurs du président'):
        ws = wb[sheet]

        cpersonne = "VIDE"
        ccabinet = "VIDE"
        cetude = "VIDE"
        corganisme = "VIDE"
        cpolitique = "VIDE"
        cparti = "VIDE"
        irow = 5
        for row in ws.iter_rows(min_row=5):
            if row[0].value:
                if row[1].value and row[0].value!=cpersonne:
                    id = row[0].value.lower()
                    cpersonne = id
                    if not id in entites.keys():
                        entites[id] = dict(nom=row[0].value,type="personne",desc="")
                else:
                    prefix = '\n' if entites[cpersonne]['desc'] else ''
                    entites[cpersonne]['desc'] += row[0].value


            if row[1].value and row[1].value!=ccabinet:
                id = normalize(row[2].value)
                ccabinet = id
                if not id in entites.keys():
                    if row[21].value:
                        periode = ' (du '+row[21].value.strftime('%d/%m/%Y') + ' au '+(row[22].value.strftime('%d/%m/%Y') if row[22].value else '?')+')'
                    else:
                        periode = ""
                    entites[id] = dict(nom=row[2].value,type="cabinet")
                liens.append(dict(source=cpersonne,target=ccabinet,type="personne-cabinet",desc="%s%s" % (row[1].value,periode)))
            if row[4].value and row[4].value!=cetude:
                id = normalize(row[4].value)
                cetude = id
                if not id in entites.keys():
                    entites[id] = dict(nom=row[4].value,type='etude')
                liens.append(dict(source=cpersonne,target=cetude,type="personne-etude",desc=row[3].value))
            if row[7].value and row[7].value!=corganisme:
                id = normalize(row[7].value)
                corganisme = id
                if not id in entites.keys():
                    entites[id] = dict(nom=row[7].value,type='organisme')
                liens.append(dict(source=cpersonne,target=corganisme,type="personne-organisme",desc=(row[6].value or '')+(" (%s)" % row[5].value if row[5].value else '')))
            if row[10].value and row[10].value!=cpolitique:
                id = normalize(row[10].value)
                cpolitique = id
                if not id in entites.keys():
                    entites[id] = dict(nom=row[10].value,type='politique')
                liens.append(dict(source=cpersonne,target=cpolitique,type="personne-politique",desc=(row[9].value or '')+(" (%s)" % row[8].value if row[8].value else '')))
            if row[11].value and row[11].value!=cparti:
                id = normalize(row[11].value)
                cparti = id
                if not id in entites.keys():
                    entites[id] = dict(nom=row[11].value,type='parti')
                liens.append(dict(source=cpersonne,target=cparti,type="personne-parti",desc=row[12].value))
            irow += 1

    return dict(entites=entites,liens=liens)
