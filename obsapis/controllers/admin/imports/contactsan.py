# -*- coding: utf-8 -*-

import requests
import re
import unicodedata

def import_contactsan_gdoc():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import requests

    r =requests.get("https://docs.google.com/spreadsheets/d/1TyyofciZokRroc2t4zAWP-abjXA-fDlG2BL3f3IVY7Y/export?format=xlsx&id=1TyyofciZokRroc2t4zAWP-abjXA-fDlG2BL3f3IVY7Y")

    f = StringIO(r.content)
    wb = load_workbook(f)

    ws = wb[wb.sheetnames[0]]
    contacts = {}
    for row in ws.iter_rows(min_row=2):
        contacts[row[0].value] = {}
        if row[2].value:
            contacts[row[0].value]['twitter']=row[2].value
        if row[3].value:
            contacts[row[0].value]['facebook']=row[3].value 

    return contacts
