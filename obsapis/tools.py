from flask import Response,request,make_response
from bson import json_util
from obsapis import use_cache,mdbrw
import datetime


def parse_content(content):
    from lxml import etree
    parser = etree.HTMLParser()
    page   = etree.fromstring(content, parser)
    return page

def maj1l(x):
    return x[0].upper()+x[1:]

def json_response(r):
    resp = Response(json_util.dumps(r))
    resp.headers['Content-Type'] = 'text/json'
    return resp

from functools import wraps
def cache_function(expires=0):
    def wrap(f):
        @wraps(f)
        def wrapped_f(*args,**kwargs):
            return use_cache(request.url,lambda:f(*args,**kwargs),expires=expires)
        return wrapped_f
    return wrap

def logitem(name,item,fields):
    def wrap(f):
        @wraps(f)
        def wrapped_f(*args,**kwargs):
            log = dict((f,request.args.get(f)) for f in fields if request.args.get(f))
            if not '127.0.0.1' in request.url and not 'api.dev' in request.url:
                log.update({ 'name':name,'item':kwargs.get(item,item),'timestamp':datetime.datetime.now(),'ip':request.environ['REMOTE_ADDR'],'user_agent':request.headers.get('User-Agent')})
                mdbrw.logs.insert_one(log)
            #print args,kwargs
            #print log
            return f(*args,**kwargs)
        return wrapped_f
    return wrap

def getdot(e,k):
    for _k in k.split('.'):
        if _k in e.keys():
            e = e[_k]
        else:
            e = ""
            break
    return e

import unicodedata
def strip_accents(s):
   return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
def normalize(s):
    return strip_accents(s).replace(u'\u2019','').replace('&apos;','').replace(u'\xa0','').encode('utf8').replace(' ','').replace("'",'').replace('-','').replace('\x0a','').replace('\xc5\x93','oe').decode('utf8').lower() if s else s


def dictToXls(data):
    from xlwt import Workbook,XFStyle
    from cStringIO import StringIO
    date_format = XFStyle()
    date_format.num_format_str = 'dd/mm/YYYY'
    stream = StringIO()
    wb = Workbook(encoding='utf8')
    from datetime import datetime,date
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.add_sheet(sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.row(0).write(j,fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldtxt = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldtxt,None)
                if isinstance(f,date) or isinstance(f,datetime):
                    ws.row(i+1).write(j,f,date_format)
                else:
                    ws.row(i+1).write(j,f)

    wb.save(stream)
    return stream.getvalue()

def dictToXlsx(data):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook


    wb = Workbook()
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.create_sheet(title=sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.cell(column=j+1,row=1,value=fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldv = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldv,None)
                ws.cell(column=j+1,row=2+i,value=f)

    return save_virtual_workbook(wb)

def xls_response(filename,v):
    import datetime
    output = make_response(v)
    output.headers['Content-Disposition'] = "attachment; filename=%s_%s.xls" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))
    output.headers['Content-type'] = 'application/vnd.ms-excel'
    return output

def image_response(type,v,filename=None,nocache=True):
    headers = {'Cache-Control':'no-cache, no-store, must-revalidate','Pragma':'no-cache'} if nocache else {}
    if filename:
        headers.update({"Content-Disposition":
                     "attachment;filename=%s-%s.png" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))})
    r = Response(v, mimetype="image/%s" % type,headers=headers)
    return r

import pyzmail
from obsapis.config_private import smtp
NOTIFY_ADDRESS = ['observatoireapi@yahoo.com']
SMTP_HOST = smtp['host']

def sendmail(sender,recipients,subject,msg='',attach=[]):
    payload, mail_from, rcpt_to, msg_id=pyzmail.compose_mail(\
        sender, \
        recipients, \
        subject, \
        'utf-8', \
        (msg, 'utf-8'), \
        html=None, \
        attachments=attach)

    #[('attached content', 'text', 'plain', 'text.txt', 'utf-8')]
    smtp_host = SMTP_HOST
    ret=pyzmail.send_mail(payload, mail_from, rcpt_to, smtp_host)

    if isinstance(ret, dict):
        if ret:
            print 'failed recipients:', ', '.join(ret.keys())
    else:
        print 'error:', ret

def api_notify(subject,msg="",attach=[],recipients=NOTIFY_ADDRESS):
    sendmail(('Observatoire API','api@observatoire-democratie.fr'),recipients,subject,msg,attach)


from bs4 import BeautifulSoup, NavigableString, Tag

def html_to_text(html):
    "Creates a formatted text email message as a string from a rendered html template (page)"
    soup = BeautifulSoup(html, 'html.parser')
    # Ignore anything in head
    body, text = soup.body, []
    for element in body.descendants:
        # We use type and not isinstance since comments, cdata, etc are subclasses that we don't want
        if type(element) == NavigableString:
            parent_tags = (t for t in element.parents if type(t) == Tag)
            hidden = False
            for parent_tag in parent_tags:
                # Ignore any text inside a non-displayed tag
                # We also behave is if scripting is enabled (noscript is ignored)
                # The list of non-displayed tags and attributes from the W3C specs:
                if (parent_tag.name in ('area', 'base', 'basefont', 'datalist', 'head', 'link',
                                        'meta', 'noembed', 'noframes', 'param', 'rp', 'script',
                                        'source', 'style', 'template', 'track', 'title', 'noscript') or
                    parent_tag.has_attr('hidden') or
                    (parent_tag.name == 'input' and parent_tag.get('type') == 'hidden')):
                    hidden = True
                    break
            if hidden:
                continue

            # remove any multiple and leading/trailing whitespace
            string = ' '.join(element.string.split())
            if string:
                if element.parent.name == 'a':
                    a_tag = element.parent
                    # replace link text with the link
                    string = a_tag['href']
                    # concatenate with any non-empty immediately previous string
                    if (    type(a_tag.previous_sibling) == NavigableString and
                            a_tag.previous_sibling.string.strip() ):
                        text[-1] = text[-1] + ' ' + string
                        continue
                elif element.previous_sibling and element.previous_sibling.name == 'a':
                    text[-1] = text[-1] + ' ' + string
                    continue
                elif element.parent.name == 'p':
                    # Add extra paragraph formatting newline
                    string = '\n' + string
                text += [string]
    doc = '\n'.join(text)
    return doc
